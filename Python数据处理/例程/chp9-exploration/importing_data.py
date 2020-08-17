"""
Part 1 / 5

Note: this is an IPython session walking through the beginning of Chapter 9 to import the
data into `agate`. It's used as an example of stepping through ideas as you work with your
data in iPython. It was saved using iPython %history -f importing_data.py so
you can use it in the repository. Obviously, if you'd like to reuse bits of this
code, it should be rewritten as a proper script :) -- @kjam
"""

import openpyxl
import agate

# 打开EXCEL文件
workbook = openpyxl.load_workbook('../../data/unicef/unicef_oct_2014.xlsx')
print(workbook.sheetnames)

# 设定活动表格，获取行数和列数
sheet = workbook.active
row = sheet.max_row
column = sheet.max_column

# 获得第5、6行的数据，并zip获得表头
rowdata5 = []
rowdata6 = []
for i in range(1, column+1):
    cellvalue = sheet.cell(row=5,column=i).value
    rowdata5.append(cellvalue)
    cellvalue = sheet.cell(row=6,column=i).value
    rowdata6.append(cellvalue)
title_rows = zip(rowdata5, rowdata6)
title_rows = list(title_rows)

titles = [t[0] + ' ' + t[1] for t in title_rows]
titles = [t.strip() for t in titles]
titles

country_rows = [sheet.cell(row=r, column=1).value for r in range(7, 114)]
print(country_rows)

from xlrd.sheet import ctype_text
import agate

text_type = agate.Text()
number_type = agate.Number()
boolean_type = agate.Boolean()
date_type = agate.Date()

example_row = sheet.row(6)
print(example_row)
print(example_row[0].ctype)
print(example_row[0].value)
print(ctype_text)


types = []

for v in example_row:
    value_type = ctype_text[v.ctype]
    if value_type == 'text':
        types.append(text_type)
    elif value_type == 'number':
        types.append(number_type)
    elif value_type == 'xldate':
        types.append(date_type)
    else:
        types.append(text_type)

types
titles

table = agate.Table(country_rows, titles, types)


def remove_bad_chars(val):
    if val == '-':
        return None
    return val

cleaned_rows = []
for row in country_rows:
    cleaned_row = [remove_bad_chars(rv) for rv in row]
    cleaned_rows.append(cleaned_row)

table = agate.Table(cleaned_rows, titles, types)
table


def get_new_array(old_array, function_to_clean):
    new_arr = []
    for row in old_array:
        cleaned_row = [function_to_clean(rv) for rv in row]
        new_arr.append(cleaned_row)
    return new_arr

cleaned_rows = get_new_array(country_rows, remove_bad_chars)

table = agate.Table(cleaned_rows, titles, types)
table.print_table(max_columns=7)
table.column_names


